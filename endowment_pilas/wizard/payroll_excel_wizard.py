import pandas as pd
import io
import os
import base64
import openpyxl
from odoo import fields, models, api, _
from odoo.exceptions import UserError
from odoo.modules import get_module_resource
from openpyxl.styles import Alignment
from datetime import datetime, date
from copy import copy

import logging

_logger = logging.getLogger(__name__)




class PayrollExcelWizard(models.TransientModel):
    _name = 'hr.payroll.excel.wizard'
    _description = 'Wizard para Reporte de Nómina a Excel'

    date_from = fields.Date(string="Fecha Desde")
    date_to = fields.Date(string="Fecha Hasta")
    payslip_run_id = fields.Many2one('hr.payslip.run', string="Lote de Nómina")

    plantilla_excel = fields.Binary(string='Plantilla Excel', help='Sube la plantilla de Excel para usar en el reporte. Si está vacío, se usará la plantilla por defecto.')
    plantilla_excel_name = fields.Char(string='Nombre Archivo Plantilla')

    #file_input = fields.Binary(string="Archivo Excel del Cliente", required=True)
    #file_name = fields.Char(string="Nombre del Archivo")



    def action_generate_excel_report(self):


        """
        Genera el reporte Excel de recibos de nómina filtrados.
        """
        # 1. DEFINICIÓN DEL DOMINIO Y BÚSQUEDA
        domain = [('state', '=', 'done')]
        if self.date_from: domain.append(('date_from', '>=', self.date_from))
        if self.date_to: domain.append(('date_to', '<=', self.date_to))
        if self.payslip_run_id: domain.append(('payslip_run_id', '=', self.payslip_run_id.id))

        payslips = self.env['hr.payslip'].search(domain)
        if not payslips:
            raise UserError(_("Nessun cedolino trovato per i filtri selezionati."))

        # 2. CARICAMENTO MODELLO (Ottimizzato per non corrompere il file)
        output = io.BytesIO()

        if self.plantilla_excel:
            # Opzione A: File caricato dall'utente
            input_buffer = io.BytesIO(base64.b64decode(self.plantilla_excel))
            wb = openpyxl.load_workbook(input_buffer, data_only=False)
        else:
            # Opzione B: File "DIVITIASSAS.xlsx" nella cartella data
            file_name_in_data = 'DIVITIASSAS.xlsx'
            path = get_module_resource('endowment_pilas', 'data', file_name_in_data)
            if not path or not os.path.exists(path):
                raise UserError(_("El modelo DIVITIASSAS.xlsx no se encuentra en la hoja de datos del módulo."))
            wb = openpyxl.load_workbook(path, data_only=False)

        # Cerchiamo il foglio "Liquidaciones" o quello attivo
        try:
            sheet = wb['Liquidaciones']
        except KeyError:
            sheet = wb.active

        # 3. LLENADO DE DATOS (Empezamos en la fila 19 porque la 17 y 18 son encabezados)
        row_num = 19
        row_counter = 1

        for payslip in payslips:
            employee = payslip.employee_id
            contract = payslip.contract_id
            correction_status_value = payslip.correction_status if hasattr(payslip, 'correction_status') else 'No'


            ############################## type identification#########################
            # Obtenemos el registro del tipo de identificación

            mapeo_id = {
                'Cédula de ciudadanía': 'CC',
                'Cédula de extranjería': 'CE',
                'Tarjeta de Identidad': 'TI',
                'Registro Civil': 'RC',
                'Pasaporte': 'PA',
                'Permiso por Protección Temporal': 'PT', # El que viste en la lista
                'PEP (Permiso Especial de Permanencia)': 'PE',
                'NIT': 'NI',
                'ID Extranjera': 'CE',
                'Documento de identificación extranjero': 'CD'
            }

            # 1. Obtenemos el registro
            tipo_doc_rec = employee.employee_address_home.l10n_latam_identification_type_id

            # 2. Extraemos el nombre (string)
            nombre_largo = tipo_doc_rec.name or ''

            # 3. Aplicamos el mapeo que definimos antes para que salga "CC", "CE", etc.
            # mapeo_id es el diccionario que definimos en el paso anterior
            tipo_doc_abreviado = mapeo_id.get(nombre_largo, nombre_largo)
            #####################################################################
            tipo_cotizante_excel_label = 'NINGUNA'
            if contract:
                if contract.pila_tipo_trabajador_id:
                    # Usamos el nombre del nuevo campo configurable, si tiene algo (como '1.Dependiente')
                    tipo_cotizante_excel_label = contract.pila_tipo_trabajador_id.name or "NINGUNA"
                #elif contract.tipo_trabajador:
                    # --- CORRECCIÓN AQUÍ ---
                    # Método recomendado para obtener la etiqueta legible del campo Selection
                    #tipo_cotizante_excel_label = dict(contract._fields['tipo_trabajador'].selection).get(contract.tipo_trabajador, '')

            sub_cotizante_excel_label = 'NINGUNA'
            if contract:
                if contract.pila_subtipo_trabajador_id:
                    sub_cotizante_excel_label = contract.pila_subtipo_trabajador_id.name or "NINGUNA"
                #elif contract.sub_tipo_trabajador:
                    # --- CORRECCIÓN AQUÍ ---
                    # Método recomendado para obtener la etiqueta legible del campo Selection
                    #sub_cotizante_excel_label = dict(contract._fields['sub_tipo_trabajador'].selection).get(contract.sub_tipo_trabajador, '')

            # --- NUEVO: Lógica para 'Horas Laboradas' ---

            dias_cotizados_pension = 0.0
            dias_cotizados_salud = 0.0
            dias_cotizados_arl = 0.0
            dias_cotizados_ccf = 0.0

            horas_laboradas = 0.0
            dias_laborados = 0.0  # 👈 nuevo acumulador

            for worked_day_line in payslip.worked_days_line_ids:
                days = worked_day_line.number_of_days
                hours = worked_day_line.number_of_hours
                code = worked_day_line.code

                # DÍAS: Se acumulan todos para que coincida con el total (sum) del XML
                dias_laborados += days

                # HORAS: Únicamente si es el código de asistencia
                if code == 'WORK100':
                    horas_laboradas += hours


                # 2. ACUMULACIÓN DE DÍAS COTIZADOS
                # Se utiliza el código de la línea del día trabajado directamente
                # para determinar a qué columna PILA se suma.

                if code == 'pension':
                    dias_cotizados_pension += days
                elif code == 'salud':
                    dias_cotizados_salud += days
                elif code == 'arl':
                    dias_cotizados_arl += days
                elif code == 'ccf':
                    dias_cotizados_ccf += days


            ####################################################

            # Inicializa ambas etiquetas
            es_extranjero_label = 'No'
            es_residente_label = 'No'

            # Lógica para 'Extranjero' (Basada en País)
            if employee and employee.country_id:
                if employee.country_id.name != 'Colombia':
                    es_extranjero_label = 'Si'

            if employee and employee.is_non_resident:
                es_residente_label = 'Si'



            # --- Lógica: Fecha de radicación en el exterior ---
            fecha_radicacion = ''
            if employee and getattr(employee, 'date_resident', False):
                fecha_radicacion = employee.date_resident  # Debe ser tipo Date en tu modelo

             # --- Lógica: Fecha inicio del contrato ---
            fecha_inicio_contrato = ''
            if contract and getattr(contract, 'date_start', False):
                fecha_inicio_contrato = contract.date_start  # Tipo Date normalmente


            # --- Lógica: ING (Ingreso) ---
            valor_ing = 'NO'
            if fecha_inicio_contrato:
                es_periodo_ingreso = False
                # Validación de rango
                if self.date_from and self.date_to:
                    if self.date_from <= fecha_inicio_contrato <= self.date_to:
                        es_periodo_ingreso = True
                elif self.date_from:
                    if fecha_inicio_contrato >= self.date_from:
                        es_periodo_ingreso = True
                elif self.date_to:
                    if fecha_inicio_contrato <= self.date_to:
                        es_periodo_ingreso = True
                else:
                    es_periodo_ingreso = True

                # Si está en el rango, aplicamos el concepto del contrato
                if es_periodo_ingreso:
                    # Buscamos el código del Many2one, si no hay, ponemos 'X'
                    valor_ing = contract.pila_ingreso_concepto_id.name or 'NO'

            # --- Lógica: Fecha final del contrato ---
            fecha_final_contrato = False
            if contract and getattr(contract, 'date_end', False):
                fecha_final_contrato = contract.date_end

            # --- Lógica: RET (Retiro) ---
            valor_ret = 'NO'
            if fecha_final_contrato:
                es_periodo_retiro = False
                # Validación de rango
                if self.date_from and self.date_to:
                    if self.date_from <= fecha_final_contrato <= self.date_to:
                        es_periodo_retiro = True
                elif self.date_from:
                    if fecha_final_contrato >= self.date_from:
                        es_periodo_retiro = True
                elif self.date_to:
                    if fecha_final_contrato <= self.date_to:
                        es_periodo_retiro = True
                else:
                    es_periodo_retiro = True

                # Si está en el rango, aplicamos el concepto del contrato
                if es_periodo_retiro:
                    # Buscamos el código del Many2one, si no hay, ponemos 'X'
                    valor_ret = contract.pila_retiro_concepto_id.name or 'NO'

            # Inicializar un diccionario para guardar los valores por defecto 'NO'
            # Asegúrate de que los códigos aquí coincidan con los de tu campo 'novelty_code'
            # 1. Inicializamos todas las novedades simples en 'NO' por defecto
            novelty_values = {code: 'NO' for code in ['TDE', 'TAE', 'TDP', 'TAP']} # Agrega aquí tus códigos

            report_start_date = self.date_from
            report_end_date = self.date_to

            afp_destino = ""
            eps_destino = ""

            # Supongamos que recorremos las administradoras configuradas en el contrato
            # o donde las tengas relacionadas (ajusta 'contract.administradora_ids' según tu modelo)
            if contract.administradoras_ids:
                for admin_line in contract.administradoras_ids:
                    # 1. ¿Es una administradora de tipo Pensión y tiene Traslado marcado?
                    if admin_line.type_entity == 'pension' and admin_line.traslado:
                        if admin_line.list_administradora_destino_id:
                            # Tomamos el nombre de la entidad destino de la lista maestra
                            afp_destino = admin_line.list_administradora_destino_id.name.upper()

                    # 2. ¿Es una administradora de tipo Salud y tiene Traslado marcado?
                    elif admin_line.type_entity == 'salud' and admin_line.traslado:
                        if admin_line.list_administradora_destino_id:
                            # Tomamos el nombre de la entidad destino de la lista maestra
                            eps_destino = admin_line.list_administradora_destino_id.name.upper()

            # Inicializar VSP con valores vacíos/por defecto
            valor_vsp = 'NO'
            fecha_vsp_str = ''

            # ----------------------------------------------------
            # 🆕 Lógica Simplificada para VSP (Variación Salario)
            # ----------------------------------------------------

            # 1. Obtenemos la fecha de cambio del campo dedicado del contrato
            fecha_cambio_sueldo = contract.date_wage_change

            if fecha_cambio_sueldo and report_start_date and report_end_date:

                # 2. Verificamos si la fecha de cambio cae DENTRO del rango del reporte.
                # El filtro que usa el reporte: [self.date_from, self.date_to]
                if report_start_date <= fecha_cambio_sueldo <= report_end_date:

                    # Si la fecha de cambio está dentro del periodo, es VSP = 'SI'
                    valor_vsp = 'SI'
                    fecha_vsp_str = fecha_cambio_sueldo # Ya es un objeto date

            #####################################################################
            # --- Lógica: VST (Variación Transitoria de Salario) ---
            valor_vst = 'NO'

            # Buscamos los payslips del empleado que estén dentro del rango y realizados (state = 'done')
            # 'contract' es el objeto del contrato actual en tu bucle
            payslips = self.env['hr.payslip'].search([
                ('employee_id', '=', contract.employee_id.id),
                ('date_from', '>=', report_start_date),
                ('date_to', '<=', report_end_date),
                ('state', '=', 'done')
            ])

            if payslips:
                # Revisamos las líneas de esas nóminas (slip_ids.line_ids)
                # Buscamos si alguna línea pertenece a una regla con is_payment_transitory = True
                # Y que el monto (total) sea mayor a cero
                transitory_lines = payslips.mapped('line_ids').filtered(
                    lambda l: l.salary_rule_id.is_payment_transitory and l.total > 0
                )

                if transitory_lines:
                    valor_vst = 'SI'

            # ----------------------------------------------------
            # 📌 Lógica para Ausencias (Usando hr.leave - Solicitudes de Ausencia)
            # ----------------------------------------------------

            # Definición de códigos relevantes y Inicialización a 'NO'
            ABSENCE_CODES = ['SLN', 'IGE', 'LMA', 'VAL-LR', 'AVP', 'VCT', 'IRL']

            # Tu búsqueda actual (sin cambios)
            leaves = self.env['hr.leave'].search([
                ('employee_id', '=', employee.id),
                ('state', '=', 'validate'),
                ('date_from', '<=', report_end_date),
                ('date_to', '>=', report_start_date),
            ])

            # Filtramos solo las que nos interesan
            relevant_leaves = []
            if leaves:
                for leave in leaves:
                    novelty_code = leave.holiday_status_id.pila_novelty_code
                    if novelty_code and novelty_code in ABSENCE_CODES:
                        relevant_leaves.append(leave)

            # Si no hay hojas relevantes, añadimos un None para que el bucle haga una iteración con todo "NO"
            if not relevant_leaves:
                relevant_leaves = [None]


            for current_leave in relevant_leaves:
                absence_novelties = {code: {'value': 'NO', 'start': '', 'end': ''} for code in ABSENCE_CODES}
                count_lma = 0
                count_ige = 0

                if current_leave:
                    novelty_code = current_leave.holiday_status_id.pila_novelty_code
                    if novelty_code == 'LMA':
                        count_lma = int(current_leave.number_of_days)
                    elif novelty_code == 'IGE':
                        count_ige = int(current_leave.number_of_days)

                    absence_novelties[novelty_code]['value'] = current_leave.holiday_status_id.name or 'SI'

                    if novelty_code != 'AVP':
                        absence_novelties[novelty_code]['start'] = current_leave.date_from.date()
                        absence_novelties[novelty_code]['end'] = current_leave.date_to.date()

                # Usaremos el nombre (label) para el Excel.
                # El método _get_selection_label() toma el nombre legible del campo de selección.
                correction_status_excel_value = payslip.correction_status if hasattr(payslip, 'correction_status') else 'no'

                # 2. Salario Mensual (Del campo 'wage' del contrato)
                salario_mensual = contract.wage if contract else 0.0

                        # ----------------------------------------------------
                # 📌 LÓGICA PARA SALARIO INTEGRAL Y VARIABLE
                # ----------------------------------------------------

                # 1. Salario Integral (True/False del contrato -> SI/NO)
                if contract and contract.wage_integral:
                    integral_excel_value = 'SI'
                else:
                    integral_excel_value = 'NO'

                # 2. Salario Variable (True/False del contrato -> SI/NO)
                if contract and contract.wage_variable:
                    variable_excel_value = 'SI'
                else:
                    variable_excel_value = 'NO'

                ##################################################################

                # ADMINISTRADORAS
                pension_admin = contract.get_admin_by_type('pension')
                salud_admin = contract.get_admin_by_type('salud')
                arl_admin = contract.get_admin_by_type('arl')
                ccf_admin = contract.get_admin_by_type('ccf')


                # --------------------------- Lógica de Alto Riesgo ---------------------------
                alto_riesgo_selection = self.env['hr.contract'].fields_get(allfields=['alto_riesgo']
                )['alto_riesgo']['selection']

                alto_riesgo_label = ''
                if contract and contract.alto_riesgo:
                    alto_riesgo_label = dict(alto_riesgo_selection).get(contract.alto_riesgo, '')


                claves_reporte = [
                    'valor_cotizacion_pension', 'valor_cotizacion_salud', 'valor_cotizacion_riesgo',
                    'valor_cotizacion_ccf', 'cotizacion_voluntaria_afiliado', 'cotizacion_voluntaria_empleador',
                    'fondo_solidaridad', 'fondo_subsistencia', 'valor_no_retenido', 'total_aportes',
                    'valor_upc', 'valor_incapacidad_eg', 'valor_licencia_maternidad',
                    'ibc','ibc_otros_parafiscales',
                    'valor_cotizacion_sena', 'valor_cotizacion_icbf', 'valor_cotizacion_esap',
                    'valor_cotizacion_men', 'exonerado_1607'
                ]

                # 2. Inicializar con 0.0 (es mejor para cálculos numéricos)
                valores_reglas = {k: 0.0 for k in claves_reporte}

                # 3. Recorrer las líneas de la nómina
                for line in payslip.line_ids:
                    # Obtenemos la marca de la regla
                    tipo = line.salary_rule_id.tipo_reporte_excel

                    # Si la regla tiene una marca y esa marca está en nuestras claves
                    if tipo and tipo in valores_reglas:
                        valores_reglas[tipo] += line.total

                ######################Otras tarifas ##################################
                admins = payslip.contract_id.administradoras_ids
                #t_ccf = sum(admins.mapped('tarifa_ccf')) or 0.0
                t_sena = sum(admins.mapped('tarifa_sena')) or 0.0
                t_icbf = sum(admins.mapped('tarifa_icbf')) or 0.0
                t_esap = sum(admins.mapped('tarifa_esap')) or 0.0
                t_men = sum(admins.mapped('tarifa_men')) or 0.0

                ################Clase ################################

                valor_clase = contract.clase if contract.clase else ''

                ################### centro trabajo ################################

                #nombre_departamento = contract.department_id.name if contract.department_id else ''
                val_centro_trabajo = ""
                if contract.work_center_id:
                    val_centro_trabajo = contract.work_center_id.name.upper()
                else:
                    # SI NO HAY REGISTRO, PASAMOS EL VALOR FIJO
                    val_centro_trabajo = "RIESGO III"

                #####################Actividad Economica #######################################
                actividad_economica = contract.economic_activitity if contract.economic_activitity else ''

                ################################upc adicional################################
                mapeo_id = {
                    'Cédula de ciudadanía': 'CC',
                    'Tarjeta de identidad': 'TI',
                    'Registro civil': 'RC',
                    'Cédula de extranjería': 'CE',
                    'Pasaporte': 'PA',
                    'NIT': 'NI',
                    'Permiso Especial de Permanencia': 'PE',
                    'Permiso por Protección Temporal': 'PT',
                }

                tipo_doc_upc_rec = employee.l10n_latam_identification_type_id

                # 2. Extraemos el nombre (o cadena vacía si no hay)
                nombre_doc_upc = tipo_doc_upc_rec.name or ''

                # 3. Buscamos la abreviatura en el mapeo. Si no está, dejamos el nombre original.
                tipo_doc_upc_abreviado = mapeo_id.get(nombre_doc_upc, nombre_doc_upc)

                # 4. Obtenemos el número y le quitamos puntos/guiones/espacios
                raw_upc = employee.upc_identification_number or ''
                numero_upc_limpio = str(raw_upc).replace('.', '').replace('-', '').strip()
                # 4. Obtenemos el número de identificación adicional
                #numero_upc = employee.upc_identification_number or ''

                data = [
                    row_counter,                                      # 1 (A)
                    tipo_doc_abreviado or 'CC',                                # 2 (B) - Debe ser 'CC', 'CE', etc.
                    str(employee.employee_address_home.vat or '').replace('.', '').replace('-', '').strip(), # 3 (C)
                    (employee.employee_address_home.last_name or '').upper(), # 4 (D)
                    (employee.employee_address_home.second_last_name or '').upper(), # 5 (E)
                    (employee.employee_address_home.first_name or '').upper(), # 6 (F)
                    (employee.employee_address_home.middle_name or '').upper(), # 7 (G)
                    (employee.employee_address_home.state_id.name or 'BOGOTA').upper(), # 8 (H)
                    (employee.employee_address_home.city_id.name or 'BOGOTA').upper(), # 9 (I)
                    tipo_cotizante_excel_label or '1. DEPENDIENTE',   # 10 (J)
                    sub_cotizante_excel_label or 'NINGUNO',           # 11 (K)
                    horas_laboradas or 0,                             # 12 (L)
                    es_extranjero_label or 'NO',                      # 13 (M)
                    es_residente_label or 'NO',                       # 14 (N)
                    fecha_radicacion or '',                           # 15 (O)

                    # --- BLOQUE DE MARCAS (SI/NO) ---
                    valor_ing or 'NO',                                # 16 (P) ING
                    fecha_inicio_contrato or '',                      # 17 (Q) Fecha ING
                    valor_ret or 'NO',                                # 18 (R) RET
                    fecha_final_contrato or '',                       # 19 (S) Fecha RET
                    novelty_values.get('TDE', 'NO'),                  # 20 (T)
                    novelty_values.get('TAE', 'NO'),                  # 21 (U)
                    novelty_values.get('TDP', 'NO'),                  # 22 (V)
                    novelty_values.get('TAP', 'NO'),                  # 23 (W)
                    valor_vsp or 'NO',                                # 24 (X) VSP
                    fecha_vsp_str or '',                              # 25 (Y) Fecha VSP
                    'NO',                                             # 26 (Z) VST (Variación Transitoria)

                    # --- BLOQUE AUSENCIAS (3 columnas por cada una) ---
                    absence_novelties['SLN']['value'], absence_novelties['SLN']['start'], absence_novelties['SLN']['end'], # 27,28,29
                    absence_novelties['IGE']['value'], absence_novelties['IGE']['start'], absence_novelties['IGE']['end'], # 30,31,32
                    absence_novelties['LMA']['value'], absence_novelties['LMA']['start'], absence_novelties['LMA']['end'], # 33,34,35
                    absence_novelties['VAL-LR']['value'], absence_novelties['VAL-LR']['start'], absence_novelties['VAL-LR']['end'], # 36,37,38
                    absence_novelties['AVP']['value'],                # 39 (Solo marca)
                    absence_novelties['VCT']['value'], absence_novelties['VCT']['start'], absence_novelties['VCT']['end'], # 40,41,42
                    absence_novelties['IRL']['value'], absence_novelties['IRL']['start'], absence_novelties['IRL']['end'], # 43,44,45

                    correction_status_excel_value or 'NO',            # 46
                    salario_mensual or 0,                             # 47
                    integral_excel_value or 'NO',                     # 48
                    variable_excel_value or 'NO',                     # 49

                    # --- PENSION ---
                    pension_admin.get_pension_label() if pension_admin else 'NINGUNO', # 50
                    dias_laborados or 0,                      # 51
                    valores_reglas['ibc'] or 0,               # 52
                    contract.get_tarifa_by_type('pension') or 0,      # 53
                    valores_reglas['valor_cotizacion_pension'] or 0,  # 54
                    alto_riesgo_label or 'NO',                        # 55
                    valores_reglas['cotizacion_voluntaria_afiliado'] or 0, # 56
                    valores_reglas['cotizacion_voluntaria_empleador'] or 0,# 57
                    valores_reglas['fondo_solidaridad'] or 0,         # 58
                    valores_reglas['fondo_subsistencia'] or 0,        # 59
                    valores_reglas['valor_no_retenido'] or 0,         # 60
                    valores_reglas['total_aportes'] or 0,             # 61
                    pension_admin.get_pension_destino_label() if pension_admin else 'NINGUNO', # 62

                    # --- SALUD ---
                    salud_admin.get_salud_label() if salud_admin else 'NINGUNO', # 63
                    dias_laborados or 0,                        # 64
                    valores_reglas['ibc'] or 0,                 # 65
                    contract.get_tarifa_by_type('salud') or 0,        # 66
                    valores_reglas['valor_cotizacion_salud'] or 0,    # 67
                    valores_reglas['valor_upc'] or 0,                 # 68
                    count_ige or '',                                  # 69 N° Autorización EG
                    valores_reglas['valor_incapacidad_eg'] or 0,      # 70
                    count_lma or '',                                  # 71 N° Autorización LMA
                    valores_reglas['valor_licencia_maternidad'] or 0, # 72
                    salud_admin.get_salud_destino_label() if salud_admin else 'NINGUNO', # 73

                    # --- RIESGOS (ARL) ---
                    arl_admin.get_arl_label() if arl_admin else 'NINGUNO', # 74
                    dias_laborados or 0,                          # 75
                    valores_reglas['ibc'] or 0,                # 76
                    contract.get_tarifa_by_type('arl') or 0,          # 77
                    valor_clase or '1',                               # 78
                    #nombre_departamento or '',                        # 79
                    val_centro_trabajo or '',                        # 79
                    actividad_economica or '',                        # 80
                    valores_reglas['valor_cotizacion_riesgo'] or 0,   # 81

                    # --- CAJA Y PARAFISCALES ---
                    dias_laborados or 0,                          # 82
                    ccf_admin.get_ccf_label() if ccf_admin else 'NINGUNO', # 83
                    valores_reglas['ibc'] or 0,                   # 84
                    contract.get_tarifa_by_type('ccf') or 0,          # 85
                    valores_reglas['valor_cotizacion_ccf'] or 0,      # 86
                    valores_reglas['ibc_otros_parafiscales'] or 0,    # 87
                    t_sena or 0,                                      # 88
                    valores_reglas['valor_cotizacion_sena'] or 0,     # 89
                    t_icbf or 0,                                      # 90
                    valores_reglas['valor_cotizacion_icbf'] or 0,     # 91
                    t_esap or 0,                                      # 92
                    valores_reglas['valor_cotizacion_esap'] or 0,     # 93
                    t_men or 0,                                       # 94
                    valores_reglas['valor_cotizacion_men'] or 0,      # 95
                    valores_reglas['exonerado_1607'] or 0,         # 96
                    tipo_doc_upc_abreviado if numero_upc_limpio else '', # 97: Solo pon tipo si hay número
                    numero_upc_limpio,
                ]

                def clean_text(val):
                    v = str(val or '').strip().upper()
                    return "" if v in ['', 'NONE', 'FALSE'] else v


                def clean_id(val):
                    v = str(val or '').replace('.', '').replace('-', '').strip()
                    return "" if v.upper() in ['', 'NONE', 'FALSE', 'NINGUNO'] else v

                def limit(val, max_len):
                    return val[:max_len] if val else ""

                # --- PROCESO DE ESCRITURA CON VALIDACIÓN DE FORMATO ---
                for col_num, cell_value in enumerate(data, start=1):
                    cell = sheet.cell(row=row_num, column=col_num)

                    # 🔍 DEBUG DE LONGITUD (AQUÍ)
                    if len(str(cell_value or '')) > 30:
                        _logger.warning(f"⚠️ FILA {row_num} COL {col_num} VALOR LARGO: {cell_value}")

                    # ---------------- COLUMNA 1 ----------------
                    # ---------------- COLUMNA 1 (A) ----------------
                    if col_num == 1:
                        cell.value = row_counter
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                        # AJUSTE DE ANCHO:
                        # 'A' es la columna 1. Un ancho de 5 a 8 es suficiente para el contador.
                        # COMENTADO PORQUE CAUSA ERROR EN APORTES EN LINEA AL CREAR COLUMNAS METADATA
                        # sheet.column_dimensions['A'].width = 8
                        continue

                    # ---------------- IDs ----------------
                    if col_num in [2, 3, 97, 98]:
                        cell.value = clean_id(cell_value)
                        cell.data_type = 's'
                        continue

                    # ---------------- NOMBRES ----------------
                    if col_num in [4, 5, 6, 7]:
                        #cell.value = clean_text(cell_value)
                        val = clean_text(cell_value)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.value = limit(val, 20)  # ajusta según PILA real
                        continue

                    # ---------------- ADMINISTRADORAS ----------------
                    if col_num in [50, 62, 63, 73, 74, 83]:
                        val = clean_text(cell_value)
                        val = val if val else "NINGUNA"
                        cell.value = limit(val, 10)
                        continue

                    # ---------------- SI / NO ----------------
                    if col_num in [48, 49, 96]:
                        val = str(cell_value or '').strip().upper()
                        cell.value = "SI" if val in ['SI', 'TRUE', '1', 'S'] else "NO"
                        continue

                    # ---------------- CLASE RIESGO ----------------
                    if col_num == 55:
                        val = str(cell_value or '').strip()
                        cell.value = "Sin Riesgo" if val in ['', '5', 'NONE', 'FALSE', '0'] else val
                        continue

                    # ---------------- CENTRO TRABAJO ----------------
                    if col_num == 79:
                        val = clean_text(cell_value)
                        val = val if val else "RIESGO III"
                        cell.value = limit(val, 10)
                        cell.data_type = 's'
                        continue

                    # ---------------- ACTIVIDAD ECONOMICA ----------------
                    if col_num == 80:
                        val = clean_text(cell_value)
                        cell.value = limit(val, 6)
                        cell.data_type = 's'
                        continue

                    # ---------------- DINERO ----------------
                    if col_num in [47, 52, 54, 56, 57, 58, 59, 60, 61, 65, 68, 70, 72, 76, 81, 84, 86, 87, 89, 91, 93, 95]:
                        try:
                            num = float(cell_value or 0)
                            if num != 0:
                                cell.value = int(num)
                                cell.number_format = '#,##0'
                            else:
                                cell.value = "" # <--- Cambiado para que quede en blanco si es 0
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        except:
                            cell.value = ""
                        continue


                    # ---------------- PORCENTAJES ----------------
                    if col_num in [53, 66, 77, 85, 88, 90, 92, 94]:
                        try:
                            val = float(str(cell_value or 0).replace('%', '').strip())
                            cell.value = val / 100 if val > 1 else val
                            cell.number_format = '0.00%'
                        except:
                            cell.value = ""
                        continue

                    # ---------------- FECHAS ----------------
                    if col_num in [15, 17, 19, 25, 28, 29, 31, 32, 34, 35, 37, 38, 41, 42, 44, 45]:
                        if cell_value:
                            try:
                                # Si ya es date/datetime, lo usamos. Si es string, lo convertimos.
                                if isinstance(cell_value, (date, datetime)):
                                    val_fecha = cell_value
                                else:
                                    val_fecha = fields.Date.from_string(cell_value)

                                cell.value = val_fecha
                                cell.number_format = 'yyyy-mm-dd' # Estándar requerido por PILA
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            except:
                                cell.value = ""
                        else:
                            cell.value = ""
                        continue

                    # ---------------- NUMÉRICOS GENERALES ----------------
                    if (47 <= col_num <= 78) or (81 <= col_num <= 95) or col_num == 12:
                        try:
                            num = float(cell_value or 0)
                            if num != 0:
                                cell.value = int(num)
                                cell.number_format = '#,##0'
                            else:
                                cell.value = ""
                        except:
                            cell.value = ""
                        continue

                    # ---------------- DEFAULT ----------------

                    # Este código solo se ejecutará si NO entró en ninguno de los IF anteriores
                    val = clean_text(cell_value)
                    cell.value = limit(val, 30)
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # --- ESTAS DOS LÍNEAS DEBEN IR AQUÍ (Fuera del bucle de columnas) ---
                row_num += 1
            row_counter += 1

        # --- TODO ESTO DEBE IR FUERA DEL BUCLE DE EMPLEADOS ---
        #last_data_row = row_num - 1

        #if sheet.max_row > last_data_row:
            #sheet.delete_rows(last_data_row + 1, sheet.max_row - last_data_row)

        # 1. Guardamos el archivo temporalmente (SIN borrar filas sobrantes)
        temp_buffer = io.BytesIO()
        wb.save(temp_buffer)
        temp_buffer.seek(0)

        # 2. Usar Pandas + XlsxWriter para generar el archivo final "Limpio"
        # Esto elimina los metadatos de dibujo corruptos que causan la reparación
        import pandas as pd
        df = pd.read_excel(temp_buffer, sheet_name=sheet.title, header=None)

        output_clean = io.BytesIO()
        writer = pd.ExcelWriter(output_clean, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Liquidaciones', index=False, header=False)

        # 3. Recuperar el libro de xlsxwriter para aplicar anchos de columna
        workbook = writer.book
        worksheet = writer.sheets['Liquidaciones']

        # Ajustamos los anchos aquí (XlsxWriter lo hace mucho mejor)
        worksheet.set_column('A:A', 10) # Columna 1
        worksheet.set_column('B:CZ', 18) # El resto

        writer.close() # Importante cerrar el writer de pandas
        out_data = output_clean.getvalue()
        output_clean.close()
        temp_buffer.close()

        # 4. Crear el adjunto
        attachment = self.env['ir.attachment'].create({
            'name': f"PILA_DIVITIASSAS_{fields.Date.today()}.xlsx",
            'type': 'binary',
            'datas': base64.b64encode(out_data), # Aquí ya tenemos los datos seguros
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # 3. Retornamos la descarga
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }



    def _format_pila(self, value, length, field_type='text'):
        """ Limpieza de datos y ajuste de ancho fijo """
        val = str(value).strip().upper() if value and not pd.isna(value) else ""

        # Novedades a veces van con 'X'
        if length == 2 and field_type == 'text' and val in ['SI', 'S', 'X', '1', 'TRUE']:
            return 'X '

        if field_type == 'num':
            # Solo números y relleno de ceros
            clean_num = "".join(filter(str.isdigit, val.split('.')[0]))
            return clean_num.zfill(length)

        if field_type == 'float_7dec':
            try:
                num = float(val)
                # La estructura pide como tarifa ej. 0.1600000
                # o el caso "0.0000000" para valores en 0
                formatted = f"{num:.7f}"
                # El usuario mando anchos de 9. "0.1600000" es len 9.
                return formatted.rjust(length)[:length]
            except:
                return "0.0000000".rjust(length)[:length]

        if field_type == 'float_5dec':
            try:
                num = float(val)
                formatted = f"{num:.5f}"
                return formatted.rjust(length)[:length]
            except:
                return "0.00000".rjust(length)[:length]

        # Texto normal: alinea a la izquierda, rellena espacios
        return val.ljust(length)[:length]


    def action_generate_txt_pila(self):
        """ Genera el Archivo Plano (TXT) con estructura Tipo 1 y Tipo 2 de PILA MinSalud """
        self.ensure_one()

        domain = [('state', '=', 'done')]
        if self.date_from: domain.append(('date_from', '>=', self.date_from))
        if self.date_to: domain.append(('date_to', '<=', self.date_to))
        if self.payslip_run_id: domain.append(('payslip_run_id', '=', self.payslip_run_id.id))

        payslips = self.env['hr.payslip'].search(domain)
        if not payslips:
            raise UserError(_("No se encontraron nóminas validadas para los filtros seleccionados."))

        company = self.env.company
        lines = []

        # =========================================================================
        # REGISTRO TIPO 1 (ENCABEZADO DE LA EMPRESA)
        # =========================================================================
        # Nota: Construimos a mano usando el formato quemado que funciona del cliente
        # para evitar fallos del T1 por ahora.
        periodo_salud = self.date_from.strftime('%Y-%m') if self.date_from else "2026-01"
        periodo_nosalud = self.date_to.strftime('%Y-%m') if self.date_to else "2025-12"

        # 011000...
        base_t1 = f"0110001{self._format_pila(company.name, 200)}S"
        base_t1 = base_t1.ljust(208)
        base_t1 += f"NI{self._format_pila(company.vat or '000000000', 16, 'num')}0E                    S54        "
        base_t1 += f"COLMENA ".ljust(40) # ARL Quemada o usar company
        base_t1 += f"14-11 {periodo_nosalud}{periodo_salud}0000000000          000440000367690240100"
        base_t1 = base_t1.ljust(305) # Ajustado al tamano del string del ejemplo T1
        lines.append(base_t1)

        # =========================================================================
        # REGISTRO TIPO 2 (DETALLE DE EMPLEADOS)
        # =========================================================================
        row_counter = 1

        for payslip in payslips:
            employee = payslip.employee_id
            contract = payslip.contract_id

            # Identificación
            mapeo_id = {
                'Cédula de ciudadanía': 'CC',
                'Cédula de extranjería': 'CE',
                'Tarjeta de Identidad': 'TI',
                'Registro Civil': 'RC',
                'Pasaporte': 'PA',
                'Permiso por Protección Temporal': 'PT',
                'PEP (Permiso Especial de Permanencia)': 'PE',
                'NIT': 'NI',
            }
            tipo_doc_rec = employee.employee_address_home.l10n_latam_identification_type_id
            tipo_doc_abreviado = mapeo_id.get(tipo_doc_rec.name or '', 'CC')
            identificacion = str(employee.employee_address_home.vat or '').replace('.', '').replace('-', '').strip()

            # Tipo cotizante (Ej: 01 Dependiente)
            tipo_cotizante = "01"
            if contract and contract.pila_tipo_trabajador_id:
                tc_str = contract.pila_tipo_trabajador_id.name
                tipo_cotizante = "".join(filter(str.isdigit, tc_str)).zfill(2)
                if not tipo_cotizante or tipo_cotizante == "00": tipo_cotizante = "01"

            # Subtipo
            sub_cotizante = "00"
            if contract and contract.pila_subtipo_trabajador_id:
                stc_str = contract.pila_subtipo_trabajador_id.name
                sub_cotizante = "".join(filter(str.isdigit, stc_str)).zfill(2)
                if not sub_cotizante: sub_cotizante = "00"

            # Nombres
            p_ape = employee.employee_address_home.last_name or ""
            s_ape = employee.employee_address_home.second_last_name or ""
            p_nom = employee.employee_address_home.first_name or ""
            s_nom = employee.employee_address_home.middle_name or ""

            salario = contract.wage if contract else 0.0

            # Novedades (ausencias)
            ABSENCE_CODES = ['SLN', 'IGE', 'LMA', 'VAL-LR', 'AVP', 'VCT', 'IRL']
            leaves = self.env['hr.leave'].search([
                ('employee_id', '=', employee.id),
                ('state', '=', 'validate'),
                ('date_from', '<=', self.date_to),
                ('date_to', '>=', self.date_from),
            ])
            novedades = {code: ' ' for code in ABSENCE_CODES}
            for leave in leaves:
                novelty_code = leave.holiday_status_id.pila_novelty_code
                if novelty_code in novedades:
                    novedades[novelty_code] = 'X'

            # Variables Numéricas
            claves_reporte = [
                'valor_cotizacion_pension', 'valor_cotizacion_salud', 'valor_cotizacion_riesgo',
                'valor_cotizacion_ccf', 'ibc', 'ibc_otros_parafiscales'
            ]
            valores_reglas = {k: 0.0 for k in claves_reporte}
            for line in payslip.line_ids:
                tipo = line.salary_rule_id.tipo_reporte_excel
                if tipo and tipo in valores_reglas:
                    valores_reglas[tipo] += line.total

            ibc = valores_reglas['ibc']

            cot_pension = valores_reglas['valor_cotizacion_pension']
            cot_salud = valores_reglas['valor_cotizacion_salud']
            cot_arl = valores_reglas['valor_cotizacion_riesgo']
            cot_caja = valores_reglas['valor_cotizacion_ccf']

            # Días cotizados
            dias_pension = 0
            dias_salud = 0
            dias_arl = 0
            dias_caja = 0
            horas_laboradas = 0

            for worked_day_line in payslip.worked_days_line_ids:
                days = int(worked_day_line.number_of_days)
                hours = int(worked_day_line.number_of_hours)
                code = worked_day_line.code
                if code == 'pension': dias_pension += days
                elif code == 'salud': dias_salud += days
                elif code == 'arl': dias_arl += days
                elif code == 'ccf': dias_caja += days

                if code == 'WORK100':
                    horas_laboradas += hours

            # Administradoras y Tarifas
            pension_admin = contract.get_admin_by_type('pension')
            cod_pension = pension_admin.code if pension_admin else "230301"
            t_pension = contract.get_tarifa_by_type('pension') or 0.16

            salud_admin = contract.get_admin_by_type('salud')
            cod_salud = salud_admin.code if salud_admin else "EPS005"
            t_salud = contract.get_tarifa_by_type('salud') or 0.125

            caja_admin = contract.get_admin_by_type('ccf')
            cod_caja = caja_admin.code if caja_admin else "CCF43"
            t_caja = contract.get_tarifa_by_type('ccf') or 0.04

            arl_admin = contract.get_admin_by_type('arl')
            cod_arl = arl_admin.code if arl_admin else ""
            t_arl = contract.get_tarifa_by_type('arl') or 0.00522

            # === INICIO CONSTUCCIÓN ESTRUCTURA TIPO 2 (Ancho Fijo) ===
            # Constrimos hasta los 693 caracteres exactos
            line2 = ""
            line2 += "02"                               # (1-2) Tipo Reg
            line2 += self._format_pila(row_counter, 5, 'num') # (3-7) Seq
            line2 += self._format_pila(tipo_doc_abreviado, 2) # (8-9) Tipo Doc
            line2 += self._format_pila(identificacion, 16)    # (10-25) Num Doc
            line2 += self._format_pila(tipo_cotizante, 2, 'num') # (26-27) Tipo Cotiz
            line2 += self._format_pila(sub_cotizante, 2, 'num') # (28-29) Subtipo
            line2 += " "                                # (30) Extranjero
            line2 += "63001"                            # (31-35) Depto/Ciudad (quemado)
            line2 += self._format_pila(p_ape, 20)       # (36-55) Ape1
            line2 += self._format_pila(s_ape, 30)       # (56-85) Ape2
            line2 += self._format_pila(p_nom, 20)       # (86-105) Nom1
            line2 += self._format_pila(s_nom, 30)       # (106-135) Nom2

            # (136-150) Banderas Novedades (ING, RET, VSP, SLN, IGE, LMA...)
            # " " o "X"
            flag_ing = "X" if getattr(contract, 'pila_ingreso_concepto_id', False) else " "
            flag_ret = "X" if getattr(contract, 'pila_retiro_concepto_id', False) else " "

            line2 += flag_ing                           # 136 ING
            line2 += flag_ret                           # 137 RET
            line2 += " "                                # 138 TDE
            line2 += " "                                # 139 TAE
            line2 += " "                                # 140 TDP
            line2 += " "                                # 141 TAP
            line2 += " "                                # 142 VSP
            line2 += " "                                # 143 VST
            line2 += novedades.get('SLN', ' ')          # 144 SLN
            line2 += novedades.get('IGE', ' ')          # 145 IGE
            line2 += novedades.get('LMA', ' ')          # 146 LMA
            line2 += novedades.get('VAL-LR', ' ')       # 147 VAC
            line2 += novedades.get('AVP', ' ')          # 148 AVP
            line2 += novedades.get('VCT', ' ')          # 149 VCT
            line2 += novedades.get('IRL', ' ')          # 150 IRL

            line2 += " 00"                              # 151-153 (Desconocido / Correcciones)

            # Códigos Administradoras (154-183)
            line2 += self._format_pila(cod_pension, 6)  # 154-159 Cod Pens
            line2 += "      "                           # 160-165 Espacio (o EPS destino)
            line2 += self._format_pila(cod_salud, 6)    # 166-171 Cod EPS
            line2 += "      "                           # 172-177 Espacio (o AFP destino)
            line2 += self._format_pila(cod_caja, 6)     # 178-183 Cod CCF

            # Días (184-191)
            line2 += self._format_pila(dias_pension, 2, 'num')  # 184-185 Dias P
            line2 += self._format_pila(dias_salud, 2, 'num')    # 186-187 Dias S
            line2 += self._format_pila(dias_arl, 2, 'num')      # 188-189 Dias A
            line2 += self._format_pila(dias_caja, 2, 'num')     # 190-191 Dias C

            # Salario e Integral
            line2 += self._format_pila(int(salario), 9, 'num')  # 192-200 Salario
            line2 += "X" if (contract and contract.wage_integral) else " "   # 201 Integral

            # IBCs
            line2 += self._format_pila(int(ibc), 9, 'num')      # 202-210 IBC P
            line2 += self._format_pila(int(ibc), 9, 'num')      # 211-219 IBC S
            line2 += self._format_pila(int(ibc), 9, 'num')      # 220-228 IBC A
            line2 += self._format_pila(int(ibc), 9, 'num')      # 229-237 IBC C

            # Pensión
            line2 += self._format_pila(t_pension, 9, 'float_7dec') # 238-246 Tarifa Pens
            line2 += self._format_pila(int(cot_pension), 9, 'num') # 247-255 Cot Pens
            line2 += self._format_pila(0, 9, 'num')               # 256-264 Aporte Sol
            line2 += self._format_pila(0, 9, 'num')               # 265-273 Aporte Sub
            line2 += self._format_pila(0, 9, 'num')               # 274-282 Val No Ret

            # Salud
            line2 += self._format_pila(t_salud, 9, 'float_7dec')  # 283-291 Tarifa Sal
            line2 += self._format_pila(int(cot_salud), 9, 'num')  # 292-300 Cot Sal
            line2 += self._format_pila(0, 9, 'num')               # 301-309 UPC

            # Incapacidades (Aut IGE, Val IGE, Aut LMA, Val LMA)
            line2 += "               "                            # 310-324 Num Aut IGE
            line2 += self._format_pila(0, 9, 'num')               # 325-333 Val IGE
            line2 += "               "                            # 334-348 Num Aut LMA
            line2 += self._format_pila(0, 9, 'num')               # 349-357 Val LMA

            # ARL
            line2 += self._format_pila(t_arl, 9, 'float_7dec')    # 358-366 Tarifa ARL
            line2 += self._format_pila("0000", 9, 'num')          # 367-375 Centro Trab
            line2 += self._format_pila(int(cot_arl), 9, 'num')    # 376-384 Cot ARL

            # CCF
            line2 += self._format_pila(t_caja, 9, 'float_7dec')   # 385-393 Tarifa CCF
            line2 += self._format_pila(int(cot_caja), 9, 'num')   # 394-402 Cot CCF

            # SENA, ICBF, ESAP, MEN
            line2 += self._format_pila(0, 9, 'float_7dec')        # 403-411 T.SENA
            line2 += self._format_pila(0, 9, 'num')               # 412-420 Cot SENA
            line2 += self._format_pila(0, 9, 'float_7dec')        # 421-429 T.ICBF
            line2 += self._format_pila(0, 9, 'num')               # 430-438 Cot ICBF
            line2 += self._format_pila(0, 9, 'float_7dec')        # 439-447 T.ESAP
            line2 += self._format_pila(0, 9, 'num')               # 448-456 Cot ESAP
            line2 += self._format_pila(0, 9, 'float_7dec')        # 457-465 T.MEN
            line2 += self._format_pila(0, 9, 'num')               # 466-474 Cot MEN

            # Adicional UPC, Exonerado, ARL
            line2 += "00"                                         # 475-476 Tipo Doc UPC
            line2 += self._format_pila("", 16)                    # 477-492 Num Doc UPC
            line2 += "S" if (valores_reglas.get('exonerado_1607', 0) > 0) else " " # 493 Exonerado
            line2 += self._format_pila(cod_arl, 6)                # 494-499 Codigo ARL
            clase = contract.clase if contract and contract.clase else "1"
            line2 += self._format_pila(clase, 1)                  # 500 Clase Riesgo
            line2 += " "                                          # 501 Ind Tarif Esp

            # BLOQUE DE FECHAS (Novedades)
            # Todo este bloque ocupa 150 caracteres en el archivo base (501 a 651)
            # Rellenamos de momento con los espacios del txt original, asegurando q mide 150
            bloque_fechas = " " * 150
            # Si hay fecha de ingreso y esta en el flag
            if flag_ing == "X" and getattr(contract, 'date_start', False):
                d_ing = contract.date_start.strftime("%Y-%m-%d")
                bloque_fechas = d_ing + bloque_fechas[10:]

            if flag_ret == "X" and getattr(contract, 'date_end', False):
                d_ret = contract.date_end.strftime("%Y-%m-%d")
                bloque_fechas = bloque_fechas[:10] + d_ret + bloque_fechas[20:]

            line2 += bloque_fechas

            # Final
            line2 += self._format_pila(0, 9, 'num')               # 652-660 IBC Otros Parafiscales
            line2 += self._format_pila(horas_laboradas, 3, 'num') # 661-663 Horas Lab
            line2 += self._format_pila("", 10)                    # 664-673 Fecha Rad Ext

            # Completamos con lo sobrante de la línea para hacer el mach exacto
            line2 = line2.ljust(693)

            lines.append(line2)
            row_counter += 1

        final_txt = "\r\n".join(lines)
        attachment_name = 'PILA_APORTES_%s.txt' % self.date_to

        attachment = self.env['ir.attachment'].create({
            'name': attachment_name,
            'type': 'binary',
            'datas': base64.b64encode(final_txt.encode('latin-1')),
            'mimetype': 'text/plain',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
